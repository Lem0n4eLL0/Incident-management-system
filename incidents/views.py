from django.shortcuts import render
from rest_framework import viewsets, permissions, filters
from .models import Incident
from .serializers import IncidentSerializer
from .filters import IncidentFilter

from rest_framework.decorators import action
from rest_framework.response import Response

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.db.models import Count
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from .permissions import IsManagerOrAdmin
from openpyxl import Workbook
from django.http import HttpResponse

from rest_framework.decorators import api_view
from rest_framework.decorators import permission_classes
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

from users.models import User

from datetime import datetime


class IsAuthorOrManager(permissions.BasePermission):
    def has_object_permission(self, request, view, obj):
        if request.user.role == 'admin':
            return True
        if request.user.role == 'manager':
            return obj.unit_snapshot == request.user.unit
        return obj.author == request.user


class IncidentViewSet(viewsets.ModelViewSet):
    serializer_class = IncidentSerializer
    filterset_class = IncidentFilter
    search_fields = ['description']

    def get_queryset(self):
        qs = Incident.objects.select_related('author', 'unit_snapshot')
        user = self.request.user
        if user.role == 'admin':
            return qs
        if user.role == 'manager':
            return qs.filter(unit_snapshot=user.unit)
        return qs.filter(author=user)

    def get_permissions(self):
        if self.request.method in ('GET', 'POST'):
            return [permissions.IsAuthenticated()]
        return [IsAuthorOrManager()]

    @action(detail=False, methods=['get'], url_path='count')
    def count(self, request):
        count = self.filter_queryset(self.get_queryset()).count()
        return Response({'count': count})


class IncidentAnalyticsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        queryset = Incident.objects.all()
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        by_status = queryset.values('status').annotate(count=Count('id'))
        by_type = queryset.values('type').annotate(count=Count('id'))
        by_unit = queryset.values('unit_snapshot__name').annotate(count=Count('id'))

        total = queryset.count()
        resolved = queryset.filter(status='завершено').count()

        return Response({
            'total': total,
            'resolved': resolved,
            'by_status': by_status,
            'by_type': by_type,
            'by_unit': by_unit,
        })


class IncidentReportXLSXAPIView(APIView):
    permission_classes = [IsAuthenticated, IsManagerOrAdmin]

    def get(self, request):
        return self._generate_report(request, request.query_params)

    def post(self, request):
        return self._generate_report(request, request.data)

    def _generate_report(self, request, params):
      #  print("params =", params)
        user = request.user

        def get_param(key):
            return params.get(key) or request.query_params.get(key)


        date_range = params.get("dateRange") or {}


        date_from = params.get("date_from") or date_range.get("from") or request.query_params.get("date_from") or request.query_params.get("dateRange.from")
        date_to = params.get("date_to") or date_range.get("to") or request.query_params.get("date_to") or request.query_params.get("dateRange.to")

        
        if date_from:
            try:
                date_from = datetime.fromisoformat(date_from.replace("Z", "")).date()
            except ValueError:
                date_from = None  

        if date_to:
            try:
                date_to = datetime.fromisoformat(date_to.replace("Z", "")).date()
            except ValueError:
                date_to = None


        status = get_param('status')
        unit = get_param('unit')
        type_filter = get_param('type')

        queryset = Incident.objects.select_related('unit_snapshot')

        
        if user.role == 'manager':
            queryset = queryset.filter(unit_snapshot=user.unit)
        elif user.role == 'admin':
            pass  
        else:
            return Response({"detail": "Forbidden"}, status=403)


        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        if status:
            queryset = queryset.filter(status=status)
        if unit:
            queryset = queryset.filter(unit_snapshot__name=unit)
        if type_filter:
            queryset = queryset.filter(type=type_filter)

        wb = Workbook()
        ws = wb.active
        ws.title = "Происшествия"

        # Стили
        header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        bold_font = Font(bold=True, color="000000")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        def apply_header_style(row_idx):
            for cell in ws[row_idx]:
                cell.fill = header_fill
                cell.font = bold_font
                cell.border = thin_border

        # Фильтры
        ws.append(["Параметры фильтрации"])
        apply_header_style(ws.max_row)

        ws.append(["Период с", date_from or "Все", "по", date_to or "Все"])
        apply_header_style(ws.max_row)

        ws.append(["Подразделение", unit or "Все"])
        apply_header_style(ws.max_row)

        ws.append(["Тип", type_filter or "Все"])
        apply_header_style(ws.max_row)

        ws.append(["Статус", status or "Все"])
        apply_header_style(ws.max_row)

        ws.append([])  # пустая строка

        
        ws.append([
            'ID', 'Номер инцидента', 'Дата', 'Тип', 'Подразделение',
            'Статус', 'Описание', 'Предпринятые меры', 'Ответственный', 'Автор'
        ])
        apply_header_style(ws.max_row)

        for i in queryset:
            ws.append([
                str(i.id),
                str(i.incident_number),
                i.date.strftime('%Y-%m-%d'),
                str(i.type),
                str(i.unit_snapshot.name if i.unit_snapshot else ''),
                str(i.status),
                str(i.description),
                str(i.measures_taken),
                str(i.responsible),
                str(i.author.full_name if i.author else '—'),
            ])

        # Автоширина
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=incidents_report.xlsx'
        wb.save(response)
        return response


class IncidentSummaryAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        today = date.today()

        
        if user.role == 'employee':
            base_qs = Incident.objects.filter(author=user)
        elif user.role == 'manager':
            base_qs = Incident.objects.filter(unit_snapshot=user.unit)
        else:  
            base_qs = Incident.objects.all()

       
        recent_incidents = base_qs.order_by('-date')[:5]

        
        current_year = today.year
        current_month = today.month
        last_year = current_year - 1
        last_month_date = today - relativedelta(months=1)

        
        total_count = base_qs.count()
        this_year_count = base_qs.filter(date__year=current_year).count()
        last_year_count = base_qs.filter(date__year=last_year).count()
        this_month_count = base_qs.filter(date__year=current_year, date__month=current_month).count()
        last_month_count = base_qs.filter(date__year=last_month_date.year, date__month=last_month_date.month).count()

        
        year_index = last_year_count - this_year_count
        month_index = last_month_count - this_month_count

        return Response({
            "recent_incidents": IncidentSerializer(recent_incidents, many=True).data,
            "stats": {
                "total": total_count,
                "year_count": this_year_count,
                "month_count": this_month_count,
                "year_index": year_index,
                "month_index": month_index
            }
        })

@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def soft_delete_incident(request, incident_id):
    try:
        incident = Incident.all_objects.get(pk=incident_id)
    except Incident.DoesNotExist:
        return Response({'detail': 'Инцидент не найден.'}, status=404)

    user = request.user

    
    if user.role != User.Role.ADMIN and incident.author != user:
        return Response({'detail': 'Недостаточно прав для удаления этого инцидента.'}, status=403)

    
    incident.is_deleted = True
    incident.save()

    
    serializer = IncidentSerializer(incident)
    return Response(serializer.data, status=200)